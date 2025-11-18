
import streamlit as st
import pdfplumber
import re
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import streamlit_authenticator as stauth

st.set_page_config(page_title="Solid Green | Net Zero (Modelled) Review", page_icon="🌿", layout="centered")

# -------------- Authentication Setup --------------
# Configure authentication
credentials = {
    'usernames': {
        'admin': {
            'name': 'Admin User',
            'password': '$2b$12$KkPOoSjUG0aU5jCv7pJGBOhebCo1RInXqAttkhHJO5A0AL486GV6C'  # Password: admin123
        },
        'reviewer': {
            'name': 'Reviewer',
            'password': '$2b$12$oqAxkd4e/U7NRTdzU.0B3OmvB94nbTggDdJHn4Ig3J2mT3OpkhH/q'  # Password: reviewer123
        },
        'Solidgreen': {
            'name': 'Solid Green',
            'password': '$2b$12$HusGyoC/XntSF4k6LPocH.vH5JRdxI00ufE3wIUf06kL2ByLXIFnG'  # Password: AaB8$5!1
        }
    }
}

authenticator = stauth.Authenticate(
    credentials,
    'solidgreen_cookie',
    'solidgreen_netze_portal_key',
    30  # cookie expiry days
)

# Login widget
authenticator.login()

# Check authentication status
if st.session_state.get('authentication_status') == False:
    st.error('Username/password is incorrect')
    st.stop()
elif st.session_state.get('authentication_status') == None:
    st.warning('Please enter your username and password')
    st.info('Demo credentials - username: `admin`, password: `admin123`')
    st.stop()

# If authenticated, show logout button below title
if st.session_state.get('authentication_status'):
    st.title("🌿 Solid Green — Net Zero Carbon (Modelled) Review Portal")
    col1, col2 = st.columns([3, 1])
    with col1:
        st.write(f"**Welcome, {st.session_state.get('name')}**")
    with col2:
        authenticator.logout(location='main')
    st.caption("Upload a Net Zero (Modelled) report PDF. Get a scored Excel checklist + infographic certificate automatically.")

# -------------- Heuristic scoring dictionary --------------
# Each item has: section, item, expected evidence (for Excel), and keywords to search for in PDF text to assign a score of 1 (strict) or 0
CHECKS = [
    # 01_General
    ("01_General","Project identification","Name, address, model type (Level 1 or 2).", [r"project:", r"location:", r"net zero carbon level", r"level\s*1|level\s*2"]),
    ("01_General","Accredited Professional","GBCSA Net Zero AP identified; signed declaration.", [r"net zero ap", r"accredited professional"]),
    ("01_General","Software","Validated simulation software (EnergyPlus/DesignBuilder/IES).", [r"energyplus", r"designbuilder", r"ies"]),
    ("01_General","Weather data","TRY/Meteonorm (≤50 km, ≥5 years).", [r"meteonorm", r"t\.?r\.?y", r"weather data"]),
    ("01_General","Emission factor","Correct national/GBCSA grid emission factor.", [r"emission factor", r"kgco2e", r"tco2e", r"grid factor"]),
    ("01_General","Modelling scope","Boundary includes all operational energy.", [r"scope", r"operational energy", r"boundary"]),
    ("01_General","Exclusions","Any exclusions justified.", [r"exclude", r"exclusion", r"not included"]),
    ("01_General","File integrity","No fatal errors; run converged.", [r"no errors", r"simulation.*(complete|converged|success)"]),
    ("01_General","Narrative summary","Short description of each major system.", [r"summary of systems", r"narrative", r"overview.*cooling", r"lighting system"]),

    # 02_Envelope_Loads
    ("02_Envelope_Loads","Wall/Roof/Floor","U-values/R-values listed.", [r"u-?value", r"r-?value", r"wall", r"roof"]),
    ("02_Envelope_Loads","Glazing","U, SHGC, VLT; frames included.", [r"glazing", r"shgc", r"vlt", r"aluminium frame|aluminum frame"]),
    ("02_Envelope_Loads","Shading","Architectural & self-shading included.", [r"shading", r"self-shading", r"overhang"]),
    ("02_Envelope_Loads","Orientation","Actual orientation modelled.", [r"orientation", r"north", r"azimuth"]),
    ("02_Envelope_Loads","Infiltration","Rate stated (e.g., 0.5 ACH).", [r"infiltration", r"ach"]),
    ("02_Envelope_Loads","Occupancy","Density & profiles.", [r"occupancy", r"m2/person|m²/person", r"schedule"]),
    ("02_Envelope_Loads","Equipment","Plug loads and schedules.", [r"plug loads", r"equipment loads", r"w/m2|w/m²"]),
    ("02_Envelope_Loads","Lighting","LPD & controls.", [r"lighting power density", r"lpd", r"occupancy sensor", r"daylight"]),
    ("02_Envelope_Loads","Hot Water","Type, capacity, COP, schedule.", [r"domestic hot water|dhw", r"heat pump", r"cop"]),
    ("02_Envelope_Loads","Vertical Transport/Ancillary","VDI 4707 or ISO 25745.", [r"vdi 4707", r"iso 25745", r"lift"]),

    # 03_HVAC_Generic
    ("03_HVAC_Generic","System overview","Cooling/heating concept described.", [r"vrf|variable refrigerant", r"chiller|chilled water", r"packaged"],),
    ("03_HVAC_Generic","Controls overview","Setpoints, zoning, scheduling.", [r"setpoint", r"schedule", r"zoning", r"econom(y|iser) cycle"]),
    ("03_HVAC_Generic","Zoning","HVAC zones align to layout.", [r"zone", r"thermal zone"]),
    ("03_HVAC_Generic","Fresh-Air Provision","Flows, fan power, control, schedules.", [r"fresh air", r"l/s", r"fan power"]),
    ("03_HVAC_Generic","Heating Method","COP>=3 or boiler efficiency; no resistive.", [r"heat pump", r"cop", r"boiler efficiency"]),
    ("03_HVAC_Generic","Fan Systems","Total fan power/SFP.", [r"sfp", r"fan power", r"specific fan power"]),
    ("03_HVAC_Generic","Thermostat Schedules","Occupied/unoccupied bands.", [r"occupied", r"unoccupied", r"thermostat"]),
    ("03_HVAC_Generic","Distribution Losses","Pumping/refrigerant losses.", [r"pumping losses", r"refrigerant losses", r"distribution losses"]),
    ("03_HVAC_Generic","VRF Outdoor Units","Capacities, COP/EER, curves.", [r"outdoor units", r"eer", r"cop", r"capacity"]),
    ("03_HVAC_Generic","VRF Indoor Units","Fan power, airflow, quantity.", [r"ceiling cassette", r"slim duct", r"fan power", r"air flow"]),
    ("03_HVAC_Generic","Chillers","Capacity/IPLV/NPLV.", [r"iplv", r"nplv", r"chiller capacity"]),
    ("03_HVAC_Generic","Cooling Towers/Pumps","Fan power, pump heads.", [r"cooling tower", r"pump head", r"pump efficiency"]),
    ("03_HVAC_Generic","Packaged/AHU DX","EER/COP, economiser logic.", [r"packaged", r"ahu", r"economiser|economizer"]),
    
    # 04_Renewables_Carbon
    ("04_Renewables_Carbon","Type & Capacity","PV type and kWp.", [r"pv", r"photovoltaic", r"kwp"]),
    ("04_Renewables_Carbon","Simulation","Yield model with losses.", [r"pvsyst", r"yield", r"performance ratio"]),
    ("04_Renewables_Carbon","Grid connection","Grid-tied/hybrid/off-grid; export/import.", [r"grid-?tied", r"hybrid", r"export", r"import"]),
    ("04_Renewables_Carbon","Metering","Meters for gen/import/export.", [r"meter", r"metering", r"ct", r"revenue meter"]),
    ("04_Renewables_Carbon","Demand vs Supply","Annual demand vs on-site gen.", [r"demand vs", r"generation", r"balance"]),
    ("04_Renewables_Carbon","Residual CO2e","Using approved factor.", [r"residual", r"tco2e", r"kgco2e"]),
    ("04_Renewables_Carbon","Offsets","Certificates (≥3 years).", [r"offset certificate", r"gold standard", r"vcs"]),
    ("04_Renewables_Carbon","Final Summary","Demand, on-site gen, offsets.", [r"summary", r"shortfall", r"offset required"]),
]

def extract_pdf_text(file_bytes):
    text = ""
    with pdfplumber.open(file_bytes) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
            text += "\n"
    text = re.sub(r"\s+", " ", text, flags=re.MULTILINE).lower()
    return text

def score_checks(text):
    results = []
    for section, item, evidence, patterns in CHECKS:
        score = 0
        for pat in patterns:
            if re.search(pat, text):
                score = 1
                break
        results.append((section, item, evidence, score))
    return results

def build_excel(results, project_name="Project", client_name="Client"):
    wb = Workbook()
    # Create sheets mapping
    def add_sheet(name):
        ws = wb.create_sheet(title=name)
        ws.append(["Section", "Check Item", "Verification Detail / Expected Evidence", "Compliance (1/0)", "Reviewer Comments"])
        # Data validation + conditional formats
        dv = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("D2:D1000")
        ws.conditional_formatting.add("D2:D1000", FormulaRule(formula=['$D2=1'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
        ws.conditional_formatting.add("D2:D1000", FormulaRule(formula=['$D2=0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
        for i,w in enumerate([24,32,70,16,48],1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    sheets = {
        "01_General": add_sheet("01_General"),
        "02_Envelope_Loads": add_sheet("02_Envelope_Loads"),
        "03_HVAC_Generic": add_sheet("03_HVAC_Generic"),
        "04_Renewables_Carbon": add_sheet("04_Renewables_Carbon"),
    }
    # Fill data per sheet
    for section, item, evidence, score in results:
        ws = sheets.get(section, None)
        if ws:
            ws.append([section, item, evidence, score, ""])

    # Results sheet
    ws_res = wb.create_sheet("05_Results_Scoring")
    ws_res.append(["Metric", "Value"])
    ws_res.append(["Items Reviewed (auto)", "=COUNTA(01_General!B2:B1000)+COUNTA(02_Envelope_Loads!B2:B1000)+COUNTA(03_HVAC_Generic!B2:B1000)+COUNTA(04_Renewables_Carbon!B2:B1000)"])
    ws_res.append(["Compliant Items (auto)", "=COUNTIF(01_General!D2:D1000,1)+COUNTIF(02_Envelope_Loads!D2:D1000,1)+COUNTIF(03_HVAC_Generic!D2:D1000,1)+COUNTIF(04_Renewables_Carbon!D2:D1000,1)"])
    ws_res.append(["Compliance % (auto)", "=(B3/B2)"])
    ws_res.append(["Outcome (auto-text)", '=IF(B4>=0.9,"Compliant","Conditionally Compliant – Revisions Required")'])
    ws_res.append(["Project", project_name])
    ws_res.append(["Client", client_name])
    for col, w in zip(["A","B"], [34, 80]):
        ws_res.column_dimensions[col].width = w

    # Remove default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def build_certificate(project, client, items_reviewed, compliant_items, score_pct, outcome_text, strengths, updates, signer="Chilufya Lombe", title="Director – Sustainability & Net Zero"):
    # PDF with ReportLab
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=60)
    title_style = ParagraphStyle('Title', fontSize=20, textColor=colors.HexColor("#006837"), alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', fontSize=13, textColor=colors.black, alignment=1, spaceAfter=18)
    section_title = ParagraphStyle('SectionTitle', fontSize=12, textColor=colors.HexColor("#006837"), spaceBefore=8, spaceAfter=6)
    normal = ParagraphStyle('Normal', fontSize=10.5, leading=15)

    elems = []
    elems.append(Paragraph("Solid Green Consulting", title_style))
    elems.append(Paragraph("<b>Net Zero Carbon (Modelled) — Outcomes Certificate</b>", subtitle_style))
    elems.append(Paragraph(f"<b>Project:</b> {project}", normal))
    if client:
        elems.append(Paragraph(f"<b>Client:</b> {client}", normal))
    elems.append(Paragraph(f"<b>Date of Review:</b> {date.today().strftime('%d %B %Y')}", normal))
    elems.append(Paragraph(f"<b>Reviewed By:</b> Solid Green Net Zero Review Engine", normal))
    elems.append(Spacer(1, 12))

    data = [
        ["Items Reviewed", str(items_reviewed)],
        ["Compliant Items", str(compliant_items)],
        ["Final Score", f"{score_pct:.1%}"],
        ["Outcome", outcome_text],
    ]
    t = Table(data, colWidths=[250, 220])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#006837")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('BOX', (0,0), (-1,-1), 1, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 14))

    elems.append(Paragraph("Key Strengths", section_title))
    elems.append(Paragraph("• " + "<br/>• ".join(strengths), normal))
    elems.append(Spacer(1, 6))

    elems.append(Paragraph("Required Updates", section_title))
    elems.append(Paragraph("• " + "<br/>• ".join(updates), normal))
    elems.append(Spacer(1, 14))

    elems.append(Paragraph("Certification Statement", section_title))
    elems.append(Paragraph(
        f"This document certifies that {project} has undergone an automated technical review for Net Zero Carbon (Modelled). "
        f"The project is expected to achieve Net Zero Carbon (Modelled) upon submission of the listed updates.", normal))
    elems.append(Spacer(1, 20))

    elems.append(Paragraph("<b>Signed on behalf of Solid Green Consulting</b>", normal))
    elems.append(Spacer(1, 8))
    elems.append(Paragraph(f"{signer}<br/>{title}<br/>Date: {date.today().strftime('%d %B %Y')}", normal))

    doc.build(elems)
    buffer.seek(0)
    return buffer

# -------------------- UI --------------------
uploaded = st.file_uploader("Upload Net Zero (Modelled) Report (PDF)", type=["pdf"])
proj = st.text_input("Project Name", value="")
client = st.text_input("Client (optional)", value="")
run = st.button("Run Automated Review")

if run:
    if not uploaded:
        st.error("Please upload a PDF report first.")
        st.stop()
    text = extract_pdf_text(uploaded)
    results = score_checks(text)

    # Excel
    excel_io = build_excel(results, project_name=proj or "Project", client_name=client or "Client")

    # Summary metrics
    total_items = len(results)
    compliant_items = sum(r[3] for r in results)
    score_pct = compliant_items / total_items if total_items else 0.0
    outcome_text = "Compliant" if score_pct >= 0.9 else "Conditionally Compliant – Revisions Required"

    # Simple strengths/updates based on high/low signals
    strengths = []
    if any(r[1]=="Wall/Roof/Floor" and r[3]==1 for r in results): strengths.append("Envelope inputs clearly documented")
    if any(r[1]=="Lighting" and r[3]==1 for r in results): strengths.append("Lighting LPD and controls well defined")
    if any(r[1]=="VRF Outdoor Units" and r[3]==1 for r in results): strengths.append("HVAC capacities and efficiencies reported")
    if any(r[1]=="Residual CO2e" and r[3]==1 for r in results): strengths.append("Carbon factor application and residual CO₂e calculated")
    if not strengths: strengths = ["Core modelling parameters identified"]

    updates = []
    if not any(r[1]=="Narrative summary" and r[3]==1 for r in results): updates.append("Add short narrative describing each major system")
    if not any(r[1]=="Controls overview" and r[3]==1 for r in results): updates.append("Describe control logic and temperature setpoints")
    if not any(r[1]=="Type & Capacity" and r[3]==1 for r in results): updates.append("Specify renewable system type and kWp")
    if not any(r[1]=="Simulation" and r[3]==1 for r in results): updates.append("Attach PV yield simulation report")
    if not any(r[1]=="Offsets" and r[3]==1 for r in results): updates.append("Attach offset certificates covering ≥3 years")
    if not updates: updates = ["Provide any missing supporting documentation referenced in the report"]

    # Certificate
    cert_io = build_certificate(
        project=proj or "Project",
        client=client or "",
        items_reviewed=total_items,
        compliant_items=compliant_items,
        score_pct=score_pct,
        outcome_text=outcome_text,
        strengths=strengths[:4],
        updates=updates[:4],
    )

    st.success("Review complete. Download your files below.")
    st.download_button("⬇️ Download Excel Checklist", data=excel_io.getvalue(), file_name=f"{proj or 'Project'}_NetZero_Modelled_Review.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.download_button("⬇️ Download Outcomes Certificate (PDF)", data=cert_io.getvalue(), file_name=f"{proj or 'Project'}_NetZero_Outcomes_Certificate.pdf", mime="application/pdf")

st.markdown("---")
st.caption("© Solid Green Consulting — Automated Net Zero (Modelled) Reviewer")
